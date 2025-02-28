import pandas as pd
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

import global_values
from base_define import CellModification, ExcelCellProc


def set_freeze(worksheet, freeze_row, freeze_col):
    freeze_cell = f"{get_column_letter(freeze_col)}{freeze_row + 1}"
    worksheet.freeze_panes = freeze_cell


def set_sorting(worksheet, sort_conditions: list):
    # 获取表头的列名，假设表头在第一行
    header_row = [cell.value for cell in worksheet[1]]

    sort_refs = []
    for col_name, order in sort_conditions:
        if col_name in header_row:
            col_idx = header_row.index(col_name) + 1  # 获取列的索引（1开始）
            sort_ref = f'{get_column_letter(col_idx)}2:{get_column_letter(col_idx)}1048576'  # 排序的范围
            sort_refs.append((sort_ref, order))

    # 设置筛选和排序条件
    worksheet.auto_filter.ref = f'A1:{get_column_letter(worksheet.max_column)}1'  # 设置自动筛选范围

    for sort_ref, order in sort_refs:
        # 添加排序条件，按升序（ascending）或降序（descending）
        worksheet.auto_filter.add_sort_condition(sort_ref, descending=order)


def post_process(original_excel_path: str, new_excel_path: str):
    workbook = load_workbook(original_excel_path)  # 直接加载Excel文件
    for metric_key in global_values.metric_key_list:
        if metric_key.excel_proc is not None:
            process_col(workbook, metric_key.metric_name, metric_key.excel_proc)

    # 1. 设置冻结窗格：冻结到第一行和C列
    set_freeze(workbook.active, 1, 4)  # 第一行冻结，C列冻结（列C的列号是3）
    # 2. 设置排序：按照列名和排序方式进行排序
    sort_conditions = [
        ("5 Yr", True),
        ("10Yr", True),
        ("3 Yr", True),
    ]
    set_sorting(workbook.active, sort_conditions)

    workbook.save(new_excel_path)  # 直接保存工作簿，修改会写入到原文件中



# process方法
def process_col(workbook, col_name: str, excel_cell_proc: ExcelCellProc) -> None:
    # 加载Excel文件进行修改（通过openpyxl）
    sheet = workbook.active  # 获取Sheet1工作表

    # 找到目标列的索引（根据列名）
    header_row = sheet[1]  # 获取第一行作为表头
    col_index = None
    for i, cell in enumerate(header_row, start=1):  # 遍历第一行的所有单元格，找到对应的列名
        if cell.value == col_name:
            col_index = i
            break

    if col_index is None:  # 如果找不到列名
        raise ValueError(f"列名 '{col_name}' 未在工作表中找到。")

    # 遍历工作表中的指定列
    for idx, row in enumerate(sheet.iter_rows(min_row=2, min_col=col_index, max_col=col_index),
                              start=2):  # 从第二行开始遍历（跳过标题）
        value = row[0].value  # 获取当前单元格的值
        if value is None:  # 如果单元格为空
            value_str: str = ''  # 将其视为空字符串
        else:
            value_str: str = str(value)  # 否则转换为字符串

        # 调用cell_func获取修改对象CellModification，传入的值是转换后的字符串
        cell_mod: CellModification = excel_cell_proc.process(value_str)

        # 获取当前单元格
        cell = sheet.cell(row=idx, column=col_index)  # 获取当前列的单元格

        # 修改单元格的内容
        if cell_mod.new_value is not None:  # 如果有新值
            cell.value = cell_mod.new_value  # 更新单元格的值

        # 修改单元格的格式（类型）
        if cell_mod.cell_type == 'number':  # 如果是数字类型
            cell.number_format = '#,##0'  # 设置数字格式
        elif cell_mod.cell_type == 'percentage':  # 如果是百分比类型
            cell.number_format = '0.00%'  # 设置为百分比格式（例如，显示为 68%）
        elif cell_mod.cell_type == 'text':  # 如果是文本类型
            cell.number_format = '@'  # 设置为纯文本格式

        # 修改单元格的背景色（如果指定了颜色）
        if cell_mod.background_color:
            fill = PatternFill(start_color=cell_mod.background_color,
                               end_color=cell_mod.background_color,
                               fill_type='solid')  # 设置背景色为指定颜色
            cell.fill = fill  # 应用背景色到单元格


if __name__ == "__main__":
    post_process(global_values.morningstar_code_excel_path, global_values.morningstar_post_process_excel_path)