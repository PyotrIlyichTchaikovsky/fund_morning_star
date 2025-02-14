import os
import re
from typing import Dict

import openpyxl
from bs4 import BeautifulSoup
from openpyxl.styles import NamedStyle
import pandas as pd


def filter_by_keyword(content: str, keyword_pattern: str) -> str:
    match = re.search(keyword_pattern, content)
    if match:
        return match.group(1)
    return "-9"


def count_keyword(content: str, target_string: str) -> str:
    count = content.count(target_string)
    if count == 0:
        count = -9
    return str(count)


def write_text_to_file(text, file_path):
    try:
        # 如果目录不存在，则创建目录
        directory = os.path.dirname(file_path)
        if not os.path.exists(directory):
            os.makedirs(directory)

        # 将文本写入文件
        with open(file_path, 'w', encoding='utf-8') as file:
            file.write(text)

        #print(f"文本已成功写入 {file_path}")
    except Exception as e:
        print(f"写入文件时发生错误: {e}")


def check_key_word_in_file(file_path: str, check_words_list: [str]) -> bool:
    if not os.path.exists(file_path):
        return False

    with open(file_path, 'r', encoding='utf-8') as file:  # 显式指定编码为utf-8
        for line in file:
            check_words_list = [check_words for check_words in check_words_list if check_words not in line]

            if len(check_words_list) == 0:
                return True

    return False


# 正则表达式：将换行符、制表符和多个空格替换成一个空格
def clean_span_content(content):
    # 替换换行符和制表符为空格
    content = re.sub(r'[\n\t\r]+', ' ', content)
    # 替换多个空格为一个空格
    content = re.sub(r'\s+', ' ', content)
    # 去除前后多余的空格
    return content.strip()


def read_from_file(disk_path: str) -> str:
    try:
        if not os.path.exists(disk_path):
            return ""

        # 将文本写入文件
        with open(disk_path, 'r', encoding='utf-8') as file:
            return file.read()
    except Exception as e:
        print(f"写入文件时发生错误: {e}")
        return ""


def try_convert_to_number(value):
    try:
        # 如果是字符串并且以百分号结尾
        if isinstance(value, str) and value.endswith('%'):
            # 替换任何非标准减号为标准减号
            value = value.replace('−', '-')  # 替换长破折号为标准减号
            # 去掉最后的 '%' 并尝试转换为浮动数字
            return float(value[:-1]) / 100

        # 尝试将值转为浮动数字
        return float(value)
    except (ValueError, TypeError):
        # 如果转换失败，返回原值
        return value


def format_columns_as_percentage(file_path, percentage_columns):
    # Load the Excel file
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    # Create a percentage style
    percent_style = NamedStyle(name="percent_style", number_format='0.00%')

    # Get the column names from the first row
    column_names = [cell.value for cell in sheet[1]]  # First row contains column headers

    # Print out the column names for debugging
    print("Column names in the file:", column_names)

    # Iterate over the specified columns
    for col in percentage_columns:
        # Check if the column exists in the sheet
        if col not in column_names:
            print(f"Warning: Column '{col}' not found in the sheet.")
            continue  # Skip this column if it's not found

        col_index = column_names.index(col) + 1  # Convert column name to index (1-based)

        # Iterate over all rows in the column (starting from the second row, excluding headers)
        for row in range(2, sheet.max_row + 1):
            cell = sheet.cell(row=row, column=col_index)

            # Check if the cell value is a valid number (and can be converted to a percentage)
            try:
                if cell.value is not None:
                    # Attempt to convert the value to a number (in case it is a string that represents a number)
                    float(cell.value)
                    cell.style = percent_style
            except (ValueError, TypeError):
                # If conversion fails (e.g., the value is not a number), leave the cell as is
                continue

    # Save the updated Excel file
    workbook.save(file_path)
    print(f"Formatted specified columns as percentage in {file_path}.")


def check_excel_for_keys(excel_path, col_name, required_keys):
    # 检查Excel文件是否存在，并且包含某些特定的key
    if not os.path.exists(excel_path):
        return False

    try:
        # 读取Excel文件
        df = pd.read_excel(excel_path)
        # 获取Excel中的所有Metric列
        metrics = df[col_name].tolist()

        # 检查是否包含所有的特定keys
        missing_keys = [key for key in required_keys if key not in metrics]

        if missing_keys:
            print(f"Missing keys: {', '.join(missing_keys)}")
            return False
        else:
            return True
    except Exception as e:
        print(f"Error reading the Excel file: {e}")
        return False


def parse_metric_from_div_pair(html_content: str) -> Dict[str, str]:
    # 解析HTML内容
    soup = BeautifulSoup(html_content, 'html.parser')

    # 初始化一个空字典
    data_dict: Dict[str, str] = {}

    # 获取所有的li标签
    items = soup.find_all('li', class_='sal-snap-panel')

    # 遍历每个li元素，提取名称和值并添加到字典中
    for item in items:
        name = item.find('div', class_='sal-dp-name').text.strip()
        value = clean_span_content(item.find('div', class_='sal-dp-value').text)
        data_dict[name] = value

    return data_dict


def parse_metric_from_table(html_content: str, table_class: str, row_key: str) -> Dict[str, str]:
    # 解析HTML内容
    soup = BeautifulSoup(html_content, 'html.parser')

    # 获取第一个符合特定class的表格
    table = soup.find('table', class_=table_class)

    if not table:
        return {}

    # 获取表头（th）的文本内容，用作字典的键
    headers = [th.get_text(strip=True) for th in table.find_all('th')]

    # 获取所有的行（tr），包括表头和数据行
    rows = table.find_all('tr')

    row_dict: Dict[str, str] = {}
    # 遍历所有数据行，查找第一列的值匹配指定值的行
    for row in rows:
        cells = row.find_all('td')
        if cells:
            first_column_value = cells[0].get_text(strip=True)
            if first_column_value == row_key:
                # 找到匹配的行，排除第一列
                for header, cell in zip(headers[1:], cells[1:]):
                    row_dict[header] = clean_span_content(cell.get_text(strip=True))
                break  # 找到对应行后停止遍历

    return row_dict


def parse_metric_from_compare_page(html_content: str) -> Dict[str, str]:
    metric_dict: dict[str, any] = {}
    soup = BeautifulSoup(html_content, 'html.parser')
    # 找到所有的<td>标签
    td_elements = soup.find_all('td')
    # 提取data-title和div的title内容的对应关系
    for idx, td in enumerate(td_elements):
        metric_name = td.get('data-title')

        div = td.find('div')
        if div:
            # 首先尝试获取div的title属性
            metric_value = div.get('title')
            # 如果div_title为空，则尝试获取div的文本内容
            if not metric_value:
                metric_value = div.get_text(strip=True)

            # 如果data_title为空，则将其修改为第几个key
            if not metric_name:
                metric_name = f"key_{idx + 1}"

            metric_dict[metric_name] = clean_span_content(metric_value)

    return metric_dict



def parse_metric_from_search_page(html_content: str) -> Dict[str, str]:
    metric_dict: dict[str, any] = {}
    soup = BeautifulSoup(html_content, 'html.parser')
    # 找到所有的<td>标签
    td_elements = soup.find_all('td')
    # 提取data-title和div的title内容的对应关系
    for idx, td in enumerate(td_elements):
        metric_name = td.get('data-title')
        if metric_name == "":
            continue

        div = td.find('div')
        if div:
            # 首先尝试获取div的title属性
            metric_value = div.get('title')
            # 如果div_title为空，则尝试获取div的文本内容
            if not metric_value:
                metric_value = div.get_text(strip=True)

            # 如果data_title为空，则将其修改为第几个key
            if not metric_name:
                metric_name = f"key_{idx + 1}"

            metric_dict[metric_name] = clean_span_content(metric_value)

    return metric_dict


def get_all_subclasses(cls):
    subclasses = cls.__subclasses__()
    result = subclasses.copy()
    for subclass in subclasses:
        result.extend(get_all_subclasses(subclass))
    return result